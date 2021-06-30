import logging
import zipfile
import os
import shutil
from lxml import etree
import openpyxl

path=os.path.dirname(os.path.abspath("x2e3.py"))
if not os.path.exists(path+"/Input") or not os.path.exists(path+"/Output"):
	os.mkdir("Input")
	os.mkdir("Output")

def extractScorm(scormName):
	fullpath = os.path.abspath(path+"/"+ scormName)
	sourceZip = zipfile.ZipFile(fullpath, 'r')
	sourceZip.extract("xml/sabaassessment.xml", path)
	return

def createTemplate(name,parsed):
	os.chdir(path+"/Template")
	template = openpyxl.load_workbook('Template.xlsx')
	os.chdir(path+"/Output")
	M = All = TF = Fill = parsed
	multiChoice = filter(lambda item: item[0]=="multiplechoicetype", M)
	allChoice = filter(lambda item: item[0]=="allthatapplytype", All)
	tfChoice = filter(lambda item: item[0]=="truefalsetype", TF)
	tfChoiceC = filter(lambda item: item[0]=="truefalsetypec", TF)
	fillerChoice = filter(lambda item: item[0]=="fillintheblanktype", Fill)

	ws1 = template.get_sheet_by_name("Multiple Choice Questions")
	ws2 = template.get_sheet_by_name("All That Apply Questions")
	ws3 = template.get_sheet_by_name("True False Questions")
	ws5 = template.get_sheet_by_name("Fill In Questions")

	row=2
	for q in multiChoice:
		row+=1
		ws1.cell(row = row, column = 1).value = q[1]
		for ans in q[2]:
			ws1.cell(row = row, column = 5).value = ans
			row+=1
		row-=len(q[2])
		for correct in q[3]:
			if correct=="true":
				ws1.cell(row = row, column = 6).value = "X"
				row+=1
			else:
				row+=1
	row=2
	for q in allChoice:
		row+=1
		ws2.cell(row = row, column = 1).value = q[1]
		for ans in q[2]:
			ws2.cell(row = row, column = 5).value = ans
			row+=1
		row-=len(q[2])
		for correct in q[3]:
			if correct=="true":
				ws2.cell(row = row, column = 6).value = "X"
				row+=1
			else:
				row+=1
	row=2
	for q in tfChoice:
		row+=1
		ws3.cell(row = row, column = 1).value = q[1]
		if q[3][0]=="true":
			ws3.cell(row = row, column = 5).value = "T"
		else:
			ws3.cell(row = row, column = 5).value = "F"

	row=2
	for q in tfChoiceC:
		row+=1
		ws3.cell(row = row, column = 1).value = q[1]
		if "true" in q[3][0]:
			ws3.cell(row = row, column = 5).value = "T"
		else:
			ws3.cell(row = row, column = 5).value = "F"
	row=2
	fillcol=5
	for q in fillerChoice:
		fillcol=5
		row+=1
		ws5.cell(row = row, column = 1).value = q[1]
		for ans in q[2]:
			if fillcol<10:
				ws5.cell(row = row, column= fillcol).value = ans
				fillcol+=1

	template.save(name+".xlsx")
	return

#replace with a function?
files = os.listdir(path)
scormList=[]
for name in files:
	if name.endswith(".zip"):
		scormList.append(name)


logging.basicConfig(filename='myapp.log', level=logging.INFO)
logging.info('Started')

for scorm in scormList:
	extractScorm(scorm)
	tree = etree.parse(path+"\\xml\sabaassessment.xml")
	root = tree.getroot()
	assessmentName = root.findtext("./AssessmentDetail/name")

	#Create a function that contains a list containg characters to remove from a string
	assessmentName=assessmentName.replace('/',' ')
	assessmentName=assessmentName.replace('\\',' ')

	questnode = root.findall(".//question")
	qlist=[]
	for node in questnode:
		q=[]
		choices=[]
		correct=[]
		q.append(node.findtext(".AssessmentQuestionDetail/question_type/name").lower())
		q.append(node.findtext(".//question_text"))
		answers = node.findall(".questionChoices//choice_text")
		anskey = node.findall(".questionChoices//correct_answer")
		for answer in answers:
			choices.append(answer.text)
		q.append(choices)
		numCorrectAns=0
		for ans in anskey:
			if ans.text =="true":
				numCorrectAns+=1
			correct.append((ans.text).lower())
		q.append(correct)
		if numCorrectAns>1 and q[0]!="fillintheblanktype":
			q[0]="allthatapplytype"
			logging.info("\n"+assessmentName+"\n"+q[1]+"\n   converted -> \"All that apply\" type.\n")
		if q[0]=="multiplechoicetype":
			if ("true" in choices[0].lower() and "false" in choices[1].lower()) or ("false" in choices[0].lower() and "true" in choices[1].lower()):
				if len(choices)==2:
					q[0]="truefalsetypec"
					logging.info("\n"+assessmentName+"\n"+q[1]+"\n   converted -> True/False type.\n")
		qlist.append(q)
	createTemplate(assessmentName, qlist)
	os.chdir(path)
	shutil.move("xml/sabaassessment.xml","Input/"+assessmentName+".xml")
	shutil.rmtree("xml")

logging.info('Finished')
#replace with a function?
"""
files = os.listdir(path)
xlsList=[]
for name in files:
	if name.endswith(".xlsx"):
		xlsList.append(name)
for xls in xlsList:
	shutil.move(xls,"Output/"+xls)
"""